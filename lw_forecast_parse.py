# -*- coding: utf-8 -*-
"""Parsuje arkusz 'Kalendarz 2026' LW -> data/lw_forecast.json (prognoza na przyszłość).
Datowane prognozy Q1-Q4 2026: okres / aktywo / sygnał / pewność / cytat LW."""
import json
import os
import re
import sys
from openpyxl import load_workbook

for s in (sys.stdout, sys.stderr):
    try: s.reconfigure(encoding="utf-8")
    except Exception: pass

SRC = r"Y:\15_AI\02_TRADING\01_MATERIALY\02_LARRY_WILLIAMS_2026\LarryWilliams_Forecast2026_Kalendarz(1).xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "lw_forecast.json")

wb = load_workbook(SRC, read_only=True, data_only=True)
ws = wb["📅 Kalendarz 2026"]
rows = [[("" if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]

items = []
quarter = ""
for r in rows:
    if not r:
        continue
    c0 = (r[0] or "").strip()
    joined = " ".join(x for x in r if x).strip()
    qm = re.search(r"Q[1-4]\s*20\d\d", joined)
    if qm and c0.startswith("──"):
        quarter = qm.group(0)
        continue
    if re.match(r"^\d+(\.\d+)?$", c0):  # wiersz prognozy
        def g(i): return (r[i].strip() if i < len(r) and r[i] else "")
        items.append({
            "q": quarter, "okres": g(1), "aktywo": g(2), "sygnal": g(3),
            "pewnosc": g(4), "wydarzenie": g(5), "ryzyko": g(7), "cytat": g(8),
        })

out = {"source": "Larry Williams Forecast 2026 — Trading Calendar (21. raport roczny)",
       "items": items}
json.dump(out, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"OK -> {OUT} | {len(items)} prognoz")
for it in items[:3]:
    print(" ", it["okres"], "|", it["aktywo"], "|", it["sygnal"])
