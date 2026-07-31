# -*- coding: utf-8 -*-
"""Parsuje kalendarz Larry'ego Williamsa (arkusz 'Oś Czasu') -> lw_seasonal.json.
Siatka AKTYWO x 12 miesiecy z sygnalem long/short/caution/neutral."""
import json
import sys
from openpyxl import load_workbook

for s in (sys.stdout, sys.stderr):
    try: s.reconfigure(encoding="utf-8")
    except Exception: pass

SRC = r"Y:\15_AI\02_TRADING\01_MATERIALY\02_LARRY_WILLIAMS_2026\LarryWilliams_Forecast2026_Kalendarz(1).xlsx"
OUT = r"Y:\15_AI\02_TRADING\sezonowosc_backtest\data\lw_seasonal.json"

wb = load_workbook(SRC, read_only=True, data_only=True)
ws = wb["📊 Oś Czasu"]
rows = [[("" if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]


def sig(v):
    v = v.upper()
    if "LONG" in v: return "long"
    if "SHORT" in v: return "short"
    if "NEUTRAL" in v: return "neutral"
    if "UWAGA" in v or "OSTRO" in v: return "caution"
    return None


def kf(name):
    u = name.upper()
    if "ZŁOTO" in u or "XAUUSD" in u or "GOLD" in u: return "gold"
    if "S&P" in u or "DJIA" in u: return "sp_djia"
    if "ROPA" in u or "XTIUSD" in u: return "oil"
    if "OBLIGACJ" in u or "TLT" in u: return "bonds"
    if u.startswith("USD") or "DOLLAR" in u: return "usd"
    if "NIERUCHOM" in u: return "reits"
    if "BITCOIN" in u or "GBTC" in u: return "btc"
    return None


hi = next(i for i, r in enumerate(rows) if r and r[0].strip().upper().startswith("AKTYWO"))
assets = {}
for r in rows[hi + 1:]:
    if not r:
        continue
    name = (r[0] or "").strip()
    if not name or name.upper().startswith("WAŻNE"):
        continue
    k = kf(name)
    if not k:
        continue
    sigs = [sig(x) for x in r[1:13]]
    if any(sigs):
        assets[k] = {"label": name, "sig": sigs}

out = {"source": "Larry Williams Forecast 2026 (Oś Czasu cykli miesięcznych, XII.2025)",
       "months": ["Sty", "Lut", "Mar", "Kwi", "Maj", "Cze", "Lip", "Sie", "Wrz", "Paź", "Lis", "Gru"],
       "assets": assets}
json.dump(out, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("OK ->", OUT)
print("assets:", list(assets.keys()))
print("gold:", assets.get("gold", {}).get("sig"))
